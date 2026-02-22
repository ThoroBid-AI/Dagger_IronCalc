#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string


ROOT = Path(__file__).resolve().parents[1]
BATCH_CSV = ROOT / "specs" / "data" / "function_batches.csv"
STATUS_CSV = ROOT / "specs" / "reports" / "oracle_capture_status.csv"


def load_batch_functions(batch_id: int) -> List[str]:
    funcs: List[str] = []
    with BATCH_CSV.open(newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if int(row["batch_id"]) == batch_id:
                funcs.append(row["function_name"])
    return funcs


def parse_batches(raw: str) -> List[int]:
    batches: List[int] = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            start_s, end_s = part.split("-", 1)
            start = int(start_s.strip())
            end = int(end_s.strip())
            if start > end:
                start, end = end, start
            batches.extend(range(start, end + 1))
        else:
            batches.append(int(part))
    seen = set()
    out = []
    for b in batches:
        if b not in seen:
            seen.add(b)
            out.append(b)
    return out


def parse_functions(raw: str) -> List[str]:
    functions: List[str] = []
    for part in raw.split(","):
        name = part.strip()
        if not name:
            continue
        functions.append(name)
    seen = set()
    out = []
    for name in functions:
        key = name.upper()
        if key not in seen:
            seen.add(key)
            out.append(name)
    return out


def sanitize_sheet_name(name: str, used: Dict[str, int]) -> str:
    safe = re.sub(r"[\[\]\*:/\\\?]", "_", name)
    safe = safe.replace("'", "")
    if len(safe) > 31:
        safe = safe[:31]
    if safe not in used:
        used[safe] = 0
        return safe
    used[safe] += 1
    suffix = f"_{used[safe]}"
    trimmed = safe[: max(0, 31 - len(suffix))]
    return f"{trimmed}{suffix}"


def parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
    if "!" in cell_ref:
        cell_ref = cell_ref.split("!")[-1]
    cell_ref = cell_ref.replace("$", "")
    col, row = coordinate_from_string(cell_ref)
    return row, column_index_from_string(col)


def iter_fixtures(fixtures_root: Path, functions: Iterable[str], engine: str):
    for func in functions:
        func_dir = fixtures_root / func
        if not func_dir.exists():
            continue
        for path in sorted(func_dir.glob("*.json")):
            data = json.loads(path.read_text())
            if data.get("engine") != engine:
                continue
            yield path, data


def format_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def batches_label(batch_ids: List[int], raw_label: str | None) -> str:
    if raw_label:
        return raw_label
    if not batch_ids:
        return ""
    if len(batch_ids) == 1:
        return str(batch_ids[0])
    return f"{batch_ids[0]}-{batch_ids[-1]}"


def update_capture_status(engine: str, batch_ids: List[int], raw_label: str | None, workbook: Path) -> None:
    STATUS_CSV.parent.mkdir(parents=True, exist_ok=True)
    label = batches_label(batch_ids, raw_label)
    rows: List[Dict[str, str]] = []
    if STATUS_CSV.exists():
        with STATUS_CSV.open(newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
    updated = False
    for row in rows:
        if row.get("engine") == engine and row.get("batches") == label:
            row["status"] = "complete"
            row["workbook"] = str(workbook)
            updated = True
            break
    if not updated:
        rows.append(
            {
                "engine": engine,
                "batches": label,
                "status": "complete",
                "workbook": str(workbook),
            }
        )
    with STATUS_CSV.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["engine", "batches", "status", "workbook"])
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Import oracle workbook results into fixtures.")
    parser.add_argument("--batch", type=int, help="Batch number to import")
    parser.add_argument(
        "--batches",
        help="Comma-separated batch list or ranges (e.g. 2-4,6). Overrides --batch.",
    )
    parser.add_argument(
        "--functions",
        help="Comma-separated function names to import directly (overrides --batches/--batch).",
    )
    parser.add_argument(
        "--label",
        help="Optional capture status label override for specs/reports/oracle_capture_status.csv.",
    )
    parser.add_argument("--engine", choices=["excel", "sheets"], required=True)
    parser.add_argument("--fixtures", default=str(ROOT / "fixtures"), help="Fixtures root")
    parser.add_argument(
        "--workbook",
        required=True,
        help="Workbook path saved from Excel/Sheets (cached results)",
    )
    args = parser.parse_args()

    batch_ids: List[int] = []
    status_label: str | None = None
    if args.functions:
        functions = parse_functions(args.functions)
        status_label = args.functions
    else:
        if args.batches:
            batch_ids = parse_batches(args.batches)
            status_label = args.batches
        elif args.batch is not None:
            batch_ids = [args.batch]
            status_label = str(args.batch)
        else:
            print("Must provide --functions or --batch/--batches")
            return 1
        functions = []
        for batch_id in batch_ids:
            functions.extend(load_batch_functions(batch_id))
    if not functions:
        if args.functions:
            print(f"No functions found for explicit function list: {args.functions}")
        else:
            print(f"No functions found for batches {batch_ids}")
        return 1

    fixtures_root = Path(args.fixtures)
    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        return 1

    wb = load_workbook(workbook_path, data_only=True)
    used_names: Dict[str, int] = {}

    updated = 0
    missing = 0

    for path, fixture in iter_fixtures(fixtures_root, functions, args.engine):
        func = fixture.get("function") or path.parent.name
        case_id = fixture.get("case_id", "case")
        sheet_name = sanitize_sheet_name(f"{func}__{case_id}", used_names)
        if sheet_name not in wb.sheetnames:
            print(f"[MISSING] Sheet {sheet_name} for {path}")
            missing += 1
            continue
        ws = wb[sheet_name]
        target = fixture.get("target") or "Sheet1!A1"
        row, col = parse_cell_ref(target)
        value = ws.cell(row=row, column=col).value
        expected_text = format_value(value)
        fixture["expected_text"] = expected_text
        if not fixture.get("notes"):
            fixture["notes"] = "Captured from oracle workbook"
        path.write_text(json.dumps(fixture, indent=2) + "\n")
        updated += 1

    print(f"Updated fixtures: {updated}")
    if missing:
        print(f"Missing sheets: {missing}")
        return 2
    if args.label:
        status_label = args.label
    update_capture_status(args.engine, batch_ids, status_label, workbook_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

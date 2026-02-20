#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string


ROOT = Path(__file__).resolve().parents[1]
BATCH_CSV = ROOT / "specs" / "data" / "function_batches.csv"


def load_batch_functions(batch_id: int) -> List[str]:
    funcs: List[str] = []
    with BATCH_CSV.open(newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if int(row["batch_id"]) == batch_id:
                funcs.append(row["function_name"])
    return funcs


def parse_batches(raw: str) -> List[int]:
    batches: List[int] = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            start_s, end_s = part.split("-", 1)
            start = int(start_s.strip())
            end = int(end_s.strip())
            if start > end:
                start, end = end, start
            batches.extend(range(start, end + 1))
        else:
            batches.append(int(part))
    # de-dupe, preserve order
    seen = set()
    out = []
    for b in batches:
        if b not in seen:
            seen.add(b)
            out.append(b)
    return out


def sanitize_sheet_name(name: str, used: Dict[str, int]) -> str:
    # Excel sheet name rules: <=31 chars, no []:*?/\
    safe = re.sub(r"[\[\]\*:/\\\?]", "_", name)
    safe = safe.replace("'", "")
    if len(safe) > 31:
        safe = safe[:31]
    if safe not in used:
        used[safe] = 0
        return safe
    used[safe] += 1
    suffix = f"_{used[safe]}"
    trimmed = safe[: max(0, 31 - len(suffix))]
    return f"{trimmed}{suffix}"


def parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
    if "!" in cell_ref:
        cell_ref = cell_ref.split("!")[-1]
    cell_ref = cell_ref.replace("$", "")
    col, row = coordinate_from_string(cell_ref)
    return row, column_index_from_string(col)


def coerce_value(value: str):
    if value == "":
        return None
    upper = value.upper()
    if upper == "TRUE":
        return True
    if upper == "FALSE":
        return False
    if re.fullmatch(r"-?\d+", value):
        try:
            return int(value)
        except Exception:
            return value
    if re.fullmatch(r"-?\d*\.\d+(?:[eE]-?\d+)?", value) or re.fullmatch(r"-?\d+(?:[eE]-?\d+)", value):
        try:
            return float(value)
        except Exception:
            return value
    return value


def iter_fixtures(fixtures_root: Path, functions: Iterable[str], engine: str):
    for func in functions:
        func_dir = fixtures_root / func
        if not func_dir.exists():
            continue
        for path in sorted(func_dir.glob("*.json")):
            data = json.loads(path.read_text())
            if data.get("engine") != engine:
                continue
            yield path, data


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate oracle capture workbook from fixtures.")
    parser.add_argument("--batch", type=int, help="Batch number to export")
    parser.add_argument(
        "--batches",
        help="Comma-separated batch list or ranges (e.g. 2-4,6). Overrides --batch.",
    )
    parser.add_argument("--engine", choices=["excel", "sheets"], required=True)
    parser.add_argument("--fixtures", default=str(ROOT / "fixtures"), help="Fixtures root")
    parser.add_argument(
        "--out",
        default=None,
        help="Output workbook path (default: fixtures/oracle_batch<B>_<engine>.xlsx)",
    )
    args = parser.parse_args()

    if args.batches:
        batch_ids = parse_batches(args.batches)
    elif args.batch is not None:
        batch_ids = [args.batch]
    else:
        print("Must provide --batch or --batches")
        return 1

    functions: List[str] = []
    for batch_id in batch_ids:
        functions.extend(load_batch_functions(batch_id))
    if not functions:
        print(f"No functions found for batches {batch_ids}")
        return 1

    fixtures_root = Path(args.fixtures)
    default_name = (
        f"oracle_batch{batch_ids[0]}_{args.engine}.xlsx"
        if len(batch_ids) == 1
        else f"oracle_batches{batch_ids[0]}-{batch_ids[-1]}_{args.engine}.xlsx"
    )
    out_path = Path(args.out) if args.out else fixtures_root / default_name

    wb = Workbook()
    index_sheet = wb.active
    index_sheet.title = "INDEX"
    index_sheet.append(["sheet", "function", "case_id", "fixture_path", "target_cell"])

    used_names: Dict[str, int] = {}

    count = 0
    for path, fixture in iter_fixtures(fixtures_root, functions, args.engine):
        func = fixture.get("function") or path.parent.name
        case_id = fixture.get("case_id", "case")
        sheet_name = sanitize_sheet_name(f"{func}__{case_id}", used_names)
        ws = wb.create_sheet(title=sheet_name)

        # inputs
        for input_item in fixture.get("inputs", []):
            cell = input_item.get("cell", "")
            value = input_item.get("value", "")
            if not cell:
                continue
            row, col = parse_cell_ref(cell)
            ws.cell(row=row, column=col, value=coerce_value(value))

        # formula
        target = fixture.get("target") or "Sheet1!A1"
        row, col = parse_cell_ref(target)
        formula = fixture.get("formula", "")
        ws.cell(row=row, column=col, value=formula)

        index_sheet.append([sheet_name, func, case_id, str(path), target])
        count += 1

    if count == 0:
        print(f"No fixtures found for batches {batch_ids} / {args.engine}")
        return 1

    wb.save(out_path)
    print(f"Wrote {out_path} ({count} fixtures)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
